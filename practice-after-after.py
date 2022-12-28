import csv
import re
import os
import numpy as np
import pandas as pd 
import pdfkit
import doctest
import cProfile
import multiprocessing
import concurrent.futures
import requests
import xmltodict
import matplotlib.pyplot as plt
from os import walk
from datetime import datetime
from openpyxl import Workbook
import multiprocessing
from prettytable import PrettyTable
from openpyxl.styles import Border, Side, Font
from jinja2 import Environment, FileSystemLoader


names = []
investment = []

"""Перевод переменной опыта в текст"""
fieldWorkExperience = {
"noExperience": "Нет опыта",
"between1And3": "От 1 года до 3 лет",
"between3And6": "От 3 до 6 лет",
"moreThan6": "Более 6 лет"
}

class AllDictionary:
    """Класс со всеми нужными словарями для вывода статистики
    Attributes:
        dictionary_salary_levels (dict): Динамика уровня зарплат по годам
        dictionary_number_vacancies (dict): Динамика количества вакансий по годам
        dictionary_salary_levels_profession (dict): Динамика уровня зарплат по годам для выбранной профессии
        dictionary_number_vacancies_profession (dict): Динамика количества вакансий по годам для выбранной профессии
        dictionary_level_salaries_cities (dict): Уровень зарплат по городам (в порядке убывания)
        dictionary_vacancy_rate_city (dict): Доля вакансий по городам (в порядке убывания)
        dictionary_currency (dict): Количество каждой валюты
        dictionary_currency_frequency (dict): Доля каждой валют 
    """
    def __init__(self):
        """Иницилизирует объект AllDictionary"""

        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        self.dictionary_currency = {}
        self.dictionary_currency_frequency = {}     

class Vacancy:
    """Класс со всеми значениями вакансий
    
    Attributes:
        name (arr): Название профессии
        area_name (arr): Компания, предоставляющая вакансию
        published_at (arr): Время публикации вакансии
        description (arr): Описание вакансии
        key_skills (arr): Навыки, которыми нужно обладать работнику
        experience_id (arr): Опыт работы
        premiumм (arr): Премиум или нет вакансия
        employer_name (arr): Название региона
        salary (class): Класс Зарплаты
    """
    def __init__(self, name, area_name, published_at, description, key_skills, experience_id, premium, employer_name, salary):
        """Иницилизирует объект Vacancy, переделывает все переменные,кроме salary в массив из одного элемента
        
        Args:
            name (str): Название профессии
            area_name (str): Компания, предоставляющая вакансию
            published_at (str): Время публикации вакансии
            description (str): Описание вакансии
            key_skills (str): Навыки, которыми нужно обладать работнику
            experience_id (str): Опыт работы
            premium (str): Премиум или нет вакансия
            employer_name (str): Название региона
            salary (str): Класс Зарплаты
        >>> Vacancy("Программист", "Компас-Плюс", "22.11.2022", "Описание", "Программирование", "between3And6", "Yes", "Новосибирск", "10000").name
        ['Программист']
        >>> type(Vacancy("Программист", "Компас-Плюс", "22.11.2022", "Описание", "Программирование", "between3And6", "Yes", "Новосибирск", "10000")).__name__
        'Vacancy'
        """
        self.name = [name]
        self.area_name = [area_name] 
        self.published_at = [published_at]
        self.description = [description]
        self.key_skills = key_skills
        self.experience_id = [experience_id]
        self.premium = [premium]
        self.employer_name = [employer_name]
        self.salary = salary

class DataSet:
    """Класс который создает и рассчитывает значения для составления таблицы и статистики
    
    Attributes:
        allDictionary (class): Класс AllDictionary
        profession_name (str): Нужная нам профессия
        profession_name (str): Название профессии, которую ищет работник
        vacancies_objects (list): Список со всеми профессиями
    """
    def __init__(self, profession = "None", fileNames = []):
        """Иницилизирует объект DataSet
        
        Args:
            profession (str): Нужная нам профессия
            fileNames (list): Массив с названиями всех файлов для считывания
        """
        self.allDictionary = AllDictionary()
        self.profession_name = profession
        self.report = Report(self.profession_name)
        self.vacancies_objects = []
        self.fileNames = fileNames

    def readystr(self, str_value):
        """Убирает в строке все лишние HTML теги
        
        Returns:
            str: Строка без HTML тегов
        
        >>> DataSet().readystr("<p>Hi</p>")
        'Hi'
        >>> DataSet().readystr("<html><head></head><p>Hi</p><body></body></html>")
        'Hi'
        """
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())
    
    def csv_ﬁlerAndReader(self, file_name):
        """Считывает файл и заполняет массив со всеми ваканскиями каждого года(каждая вакансяи с 9 полями, которых нет-делает пустыми)
        
        Args:
            file_name (str): Название файла с профессиями по годам
        Returns:
            list: Список со всеми вакансиями
        """
        names = []
        investment = []
        vacancies_objects = []
        with open("newCSV/" + file_name, encoding="UTF-8-sig") as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                if(len(names) == 0):
                    names = row
                elif(len(names) == len(row) and not("" in row)):
                    investment.append(row)

            for element in investment:
                array = ["","","","","","","","",""]
                namesIndex = ["name", "area_name", "published_at", "description", "key_skills", "experience_id", "premium", "employer_name", "salary"]
                for item in range(len(names)):
                    value = element[item]
                    value = value.split("\n") if ("\n" in value) else [value]
                    arrayWithoutTegs = []
                    for word in value:
                        if(word.upper() == "TRUE"):
                            arrayWithoutTegs.append("Да")
                        elif(word.upper() == "FALSE"):
                            arrayWithoutTegs.append("Нет")
                        elif(word in fieldWorkExperience.keys()):
                            arrayWithoutTegs.append(fieldWorkExperience[word])
                        else:
                            arrayWithoutTegs.append(self.readystr(word))
                    if(len(arrayWithoutTegs) == 1):
                        arrayWithoutTegs = arrayWithoutTegs[0]
                    array[namesIndex.index(names[item])] = arrayWithoutTegs
                vacancy = Vacancy(*array)
                vacancies_objects.append(vacancy)
            listWithSumSalaryAndCount = self.countsNumberAndSalary({vacancy.published_at[0][0:4]:vacancies_objects})
            return listWithSumSalaryAndCount

    def runningFunctionsInMultiThread(self, fileNames):
        """Функция запускает другие функции в многопотоке

        Args:
            fileNames (list): Список с названиями всех файлов csv
        Returns:
            listWithAllVacancies (list): Список всех вакансий
        """
        listWithAllVacancies = []
        with concurrent.futures.ProcessPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(self.csv_ﬁlerAndReader, fileName): fileName for fileName in fileNames}
            for fut in concurrent.futures.as_completed(futures):
                listWithAllVacancies.append(fut.result())
        return listWithAllVacancies

    def splitFileByYear(self, fileName):
        """Функция разбивает один большой файл на файлы по годам публикации вакансий и записывает их в указазнную папку(newCSV)

        Args:
            fileName (str): Название файла, который нужно разбить
        """
        dictVithVacancies = {}
        def write_chunk(part, lines):
            with open('newCSV//data_part_'+ str(part) +'.csv', 'w', encoding="utf-8-sig") as f_out:
                f_out.writelines(lines)
                f_out.close()
        
        with open(fileName, 'r', encoding="utf-8-sig") as fi_le:
            names = fi_le.readline()
            for string in fi_le:
                year = string.split(",")[len(string.split(",")) - 1][0:4]
                if(year in dictVithVacancies.keys()):
                    dictVithVacancies[year].append(string)
                else: 
                    dictVithVacancies[year] = [names, string]

            if len(dictVithVacancies) > 0:
                for vacancie in dictVithVacancies:
                    write_chunk(vacancie, dictVithVacancies[vacancie])

    def countsNumberAndSalary(self, everyYear):
        """Функция считает количество ваканский в каждом году, сумму всех зарплат, количество ваканский(которую просят) в каждом году,
        сумму зарплат ваканский(которую просят) в каждом году
        
        Args:
            everyYear (dict): Словарь с ключом - год, а значение - список всех вакансий
        Returns:
            list: Возвращает список из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        sumAverageSalary = 0

        countAverageSalaryForProfession = 0
        sumAverageSalaryForProfession = 0
        for vacancie in list(everyYear.values())[0]:
            sumAverageSalary += float(vacancie.salary)
            if(self.profession_name in vacancie.name[0]):
                countAverageSalaryForProfession += 1
                sumAverageSalaryForProfession += float(vacancie.salary)
        return [list(everyYear.keys())[0],len(list(everyYear.values())[0]), sumAverageSalary, countAverageSalaryForProfession, sumAverageSalaryForProfession, list(everyYear.values())[0]]   

    def readyPrint(self, listWithSumSalaryAndCount):
        """Функция сначала считает заполняет словари с городами,
        потом выводит всю статистику, формирует графики,excel,pdf файлы
        
        Args:
            listWithSumSalaryAndCount (list): Cписок из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        for statisticForYear in listWithSumSalaryAndCount:
            self.vacancies_objects += statisticForYear[5]
            if(statisticForYear[0] in self.allDictionary.dictionary_salary_levels):
                #1 заполнения словаря Динамика уровня зарплат по годам
                self.allDictionary.dictionary_salary_levels[statisticForYear[0]] += int(statisticForYear[2] / statisticForYear[1])
                self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
                #2 заполнения словаря Динамика количества вакансий по годам
                self.allDictionary.dictionary_number_vacancies[statisticForYear[0]] += statisticForYear[1]
                self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
                #3 заполнения словаря Динамика уровня зарплат по годам для выбранной профессии
                self.allDictionary.dictionary_salary_levels_profession[statisticForYear[0]] += int(statisticForYear[4] / statisticForYear[3])
                self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
                #4 заполнения словаря Динамика количества вакансий по годам для выбранной профессии
                self.allDictionary.dictionary_number_vacancies_profession[statisticForYear[0]] += statisticForYear[3]
                self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession
            else:
                #1 заполнения словаря Динамика уровня зарплат по годам
                self.allDictionary.dictionary_salary_levels[statisticForYear[0]] = int(statisticForYear[2] / statisticForYear[1])
                self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
                #2 заполнения словаря Динамика количества вакансий по годам
                self.allDictionary.dictionary_number_vacancies[statisticForYear[0]] = statisticForYear[1]
                self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
                #3 заполнения словаря Динамика уровня зарплат по годам для выбранной профессии
                self.allDictionary.dictionary_salary_levels_profession[statisticForYear[0]] = int(statisticForYear[4] / statisticForYear[3])
                self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
                #4 заполнения словаря Динамика количества вакансий по годам для выбранной профессии
                self.allDictionary.dictionary_number_vacancies_profession[statisticForYear[0]] = statisticForYear[3]
                self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession

        print("Динамика уровня зарплат по годам: ",end="")
        self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels = dict(sorted(self.allDictionary.dictionary_salary_levels.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_salary_levels)
        print("Динамика количества вакансий по годам: ",end="")
        self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies = dict(sorted(self.allDictionary.dictionary_number_vacancies.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_number_vacancies)
        print("Динамика уровня зарплат по годам для выбранной профессии: ",end="")
        self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession = dict(sorted(self.allDictionary.dictionary_salary_levels_profession.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_salary_levels_profession)
        print("Динамика количества вакансий по годам для выбранной профессии: ",end="")
        self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession = dict(sorted(self.allDictionary.dictionary_number_vacancies_profession.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_number_vacancies_profession)  
        
        self.report.generate_graphics()
        self.report.generate_excel()
        self.report.generate_pdf()
        
    def calculating_average_salary(self):
        """Функция считает значения среднии зарплаты в разных городах, долю ваканский в каждом городе относительно всех вакансий"""
        dicti3 = self.allDictionary.dictionary_level_salaries_cities
        for salaryYear in dicti3.keys():
            dicti3[salaryYear] = round(float(float(dicti3[salaryYear].split(" ")[1]) / float(dicti3[salaryYear].split(" ")[0])), 2)
        dicti4 = self.allDictionary.dictionary_vacancy_rate_city
        for salaryYear in dicti4.keys(): 
            dicti4[salaryYear] = round(dicti4[salaryYear] / len(self.vacancies_objects), 4)
    
class Report:
    """Класс который создает и рассчитывает значения для составления таблицы и статистики
    
    Attributes:
        dictionary_salary_levels (dict): Динамика уровня зарплат по годам
        dictionary_number_vacancies (dict): Динамика количества вакансий по годам
        dictionary_salary_levels_profession (dict): Динамика уровня зарплат по годам для выбранной профессии
        dictionary_number_vacancies_profession (dict): Динамика количества вакансий по годам для выбранной профессии
        dictionary_level_salaries_cities (dict): Уровень зарплат по городам (в порядке убывания)
        dictionary_vacancy_rate_city (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, profession):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        self.profession = profession

    def generate_graphics(self):
        """Генерация всех графиков в картинку при помоще библиотеки matplotlib"""
        plt.rcParams.update({'font.size': 8})
        #1 ГРАФИК
        labelsLevel = self.dictionary_salary_levels.keys()
        middle_salary = self.dictionary_salary_levels.values()
        middle_salary_profession = self.dictionary_salary_levels_profession.values()

        x = np.arange(len(labelsLevel))  
        width = 0.35 

        ax = plt.subplot(221)
        ax.bar(x - width/2, middle_salary, width, label='средняя з/п')
        ax.bar(x + width/2, middle_salary_profession, width, label='з/п ' + self.profession)
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(x, labelsLevel, rotation = 90)
        plt.grid(axis='y')
        ax.legend()
        #2 ГРАФИК
        labelsCount = self.dictionary_number_vacancies.keys()
        middle_count = self.dictionary_number_vacancies.values()
        middle_count_profession = self.dictionary_number_vacancies_profession.values()

        x = np.arange(len(labelsCount))  
        width = 0.35 

        ax = plt.subplot(222)
        ax.bar(x - width/2, middle_count, width, label='Количества вакансий')
        ax.bar(x + width/2, middle_count_profession, width, label='Количества вакансий \n' + self.profession)
        ax.set_title('Количество вакансий по годам')
        ax.set_xticks(x, labelsCount, rotation = 90)
        plt.grid(axis='y')
        ax.legend()

        plt.subplots_adjust(wspace = 0.5, hspace = 0.5)
        plt.savefig("Graphics", dpi = 200, bbox_inches='tight')

    def set_border(self, ws, cell_range):
        """Генерация всех графиков в картинку
        
        Args:
            ws (worksheet): рабочая страница в Excel 
            cell_range (str): строка, в которой переданы ячейки, которые нудно изменить 
        """
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)    

    def set_bold(self, ws, cell_range):
        """Делает жирный шрифт в шапке таблицы
        Args:
            ws (worksheet): рабочая страница в Excel 
            cell_range (str): строка, в которой переданы ячейки, которые нудно изменить    
        """
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold = True)

    def width(self, ws):
        """Задаёт размеры каждого столбца в таблице
        
        Args:
            ws (worksheet): рабочая страница в Excel 
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.5
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate_excel(self):
        """Генерирует таблицу Excel"""
        wb = Workbook()
        #1 СТРАНИЦА
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws1.append(["Год", "Средняя зарплата", "Средняя зарплата - " + self.profession, "Количество вакансий", "Количество вакансий - " + self.profession])
        
        countIndexA = 2
        for year in self.dictionary_salary_levels.keys():
            ws1['A' + str(countIndexA)] = year
            countIndexA += 1
        countIndexB = 2
        for value in self.dictionary_salary_levels.values():
            ws1['B' + str(countIndexB)] = value
            countIndexB += 1
        countIndexC = 2
        for value in self.dictionary_salary_levels_profession.values():
            ws1['C' + str(countIndexC)] = value
            countIndexC += 1
        countIndexD = 2
        for value in self.dictionary_number_vacancies.values():
            ws1['D' + str(countIndexD)] = value
            countIndexD += 1
        countIndexE = 2
        for value in self.dictionary_number_vacancies_profession.values():
            ws1['E' + str(countIndexE)] = value
            countIndexE += 1
        
        self.width(ws1)
        self.set_border(ws1, 'A1:E' + str(1 + len(self.dictionary_salary_levels.keys()))) 
        self.set_bold(ws1, 'A1:E1') 
        wb.save('report.xlsx')

    def generate_pdf(self):
        """Генерирует файл PDF"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")
        
        pdf_template = template.render({"name": self.profession})
        config = pdfkit.configuration(wkhtmltopdf = r'C:\Users\Giga\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf_template = pdf_template.replace("path", os.path.abspath(os.curdir))
        options = {'enable-local-file-access': None}

        readyHtmlTable = self.generate_table()
        pdf_template = pdf_template.replace("$table1;", readyHtmlTable)
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def generate_table(self):
        """Создаёт таблицу для HTML кода
        
        >>> profession = "Программист"
        >>> Report(profession).generate_table()
        "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - Программист</th><th>Количество вакансий</th><th>Количество вакансий - Программист</th></tr></tr></table>"
        """
        #1 ТАБЛИЦА
        readyHtmlTable = "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - "
        readyHtmlTable += self.profession + "</th><th>Количество вакансий</th><th>Количество вакансий - " + self.profession + "</th></tr>"
        for index in range(len(list(self.dictionary_salary_levels.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels_profession.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies_profession.values())[index]) + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table>"
        return readyHtmlTable

def main():
    #folderName = input("Введите название папки: ")
    fileName = input("Введите название файла: ")
    profession = input("Введите название профессии: ")
    folderName = "newCSV"
    #fileName = "newDataFrame.csv"
    #profession = "Программист"
    
    fileNames = [] 


    for (dirpath, dirnames, filenames) in walk(folderName):
        fileNames.extend(filenames)
        break

    printer = DataSet(profession, fileNames)
    printer.splitFileByYear(fileName)
    listWithSumSalaryAndCount = printer.runningFunctionsInMultiThread(fileNames)   
    printer.readyPrint(listWithSumSalaryAndCount)


if __name__ == "__main__":
    doctest.testmod()
    cProfile.run('main()')