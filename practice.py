import csv
import re
import os
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
import doctest
import cProfile
import multiprocessing
import concurrent.futures
from os import walk
from datetime import datetime
from openpyxl import Workbook
import multiprocessing
from prettytable import PrettyTable
from openpyxl.styles import Border, Side, Font
from jinja2 import Environment, FileSystemLoader


names = []
investment = []

"""Словарь по переводу валюты в рубли"""
currency_to_rub = {  
    "AZN": 35.68,  
    "BYR": 23.91,  
    "EUR": 59.90,  
    "GEL": 21.74,  
    "KGS": 0.76,  
    "KZT": 0.13,  
    "RUR": 1,  
    "UAH": 1.64,  
    "USD": 60.66,  
    "UZS": 0.0055,  
}

"""Перевод переменных на анлийский язык"""
fieldToEng = {
"Название": "name",
"Описание":"description",
"Навыки":"key_skills",
"Опыт работы":"experience_id",
"Премиум-вакансия":"premium",
"Компания":"employer_name",
"Оклад": "currency",
"Название региона":"area_name",
"Дата публикации вакансии":"published_at",
"Идентификатор валюты оклада":"currency"
}

"""Перевод переменных на русский язык"""
fieldToRus = {
"name":"Название",
"description":"Описание",
"key_skills":"Навыки",
"experience_id":"Опыт работы",
"premium":"Премиум-вакансия",
"employer_name":"Компания",
"currency":"Оклад",
"area_name":"Название региона",
"published_at":"Дата публикации вакансии"
}

"""Расшифрования сокращения валют"""
fieldCurrency = {
"AZN": "Манаты",
"BYR": "Белорусские рубли",
"EUR": "Евро",
"GEL": "Грузинский лари",
"KGS": "Киргизский сом",
"KZT": "Тенге",
"RUR": "Рубли",
"UAH": "Гривны",
"USD": "Доллары",
"UZS": "Узбекский сум"
}

"""Перевод переменной опыта в текст"""
fieldWorkExperience = {
"noExperience": "Нет опыта",
"between1And3": "От 1 года до 3 лет",
"between3And6": "От 3 до 6 лет",
"moreThan6": "Более 6 лет"
}

class AllDictionary:
    """Класс со всеми нужными словарями для вывода статистики
    Attributes:
        dictionary_salary_levels (dict): Динамика уровня зарплат по годам
        dictionary_number_vacancies (dict): Динамика количества вакансий по годам
        dictionary_salary_levels_profession (dict): Динамика уровня зарплат по годам для выбранной профессии
        dictionary_number_vacancies_profession (dict): Динамика количества вакансий по годам для выбранной профессии
        dictionary_level_salaries_cities (dict): Уровень зарплат по городам (в порядке убывания)
        dictionary_vacancy_rate_city (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self):
        """Иницилизирует объект AllDictionary"""

        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        
class Salary:
    """Класс со всеми значениями зарплаты
    
    Attributes:
        salary_from (list): Минимальное значение зарплаты
        salary_to (list): Максимальное значение зарплаты
        salary_gross (list): Наличие или отсутсвие включенных налогов
        salary_currency (list): Валюта в которой будет выплачиваться зарплата
        salary (str): Красивый вывод зарплаты
    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Иницилизирует объект Salary, переделывает все переменные,кроме salary в массив из одного элемента
        Args:
            salary_from (str): Минимальное значение зарплаты
            salary_to (str): Максимальное значение зарплаты
            salary_gross (str): Наличие или отсутсвие включенных налогов
            salary_currency (str): Валюта в которой будет выплачиваться зарплата
        
        >>> Salary(10000, 20000, "Yes", 'EUR').salary_from
        [10000]
        >>> Salary(10000, 20000.0, "Yes", 'EUR').salary
        '10 000 - 20 000 (Евро) (С вычетом налогов)'
        >>> type(Salary(10000, 20000.0, "Yes", 'EUR')).__name__
        'Salary'
        """
        self.salary_from = [salary_from]
        self.salary_to = [salary_to]
        self.salary_gross = [salary_gross]
        self.salary_currency = [salary_currency]
        self.salary = str('{:,}'.format(int(float(salary_from))).replace(',', ' ')) + " - " + str('{:,}'.format(int(float(salary_to))).replace(',', ' ')) + " (" + fieldCurrency[salary_currency] + ") (" + ("Без вычета налогов" if salary_gross.upper() == "ДА" else "С вычетом налогов") + ")"

class Vacancy:
    """Класс со всеми значениями зарплаты
    
    Attributes:
        name (arr): Название профессии
        area_name (arr): Компания, предоставляющая вакансию
        published_at (arr): Время публикации вакансии
        description (arr): Описание вакансии
        key_skills (arr): Навыки, которыми нужно обладать работнику
        experience_id (arr): Опыт работы
        premiumм (arr): Премиум или нет вакансия
        employer_name (arr): Название региона
        salary (class): Класс Зарплаты
        elements (arr): Все значение в одной переменной
    """
    def __init__(self, name, area_name, published_at, description, key_skills, experience_id, premium, employer_name, salary):
        """Иницилизирует объект Vacancy, переделывает все переменные,кроме salary в массив из одного элемента
        
        Args:
            name (str): Название профессии
            area_name (str): Компания, предоставляющая вакансию
            published_at (str): Время публикации вакансии
            description (str): Описание вакансии
            key_skills (str): Навыки, которыми нужно обладать работнику
            experience_id (str): Опыт работы
            premium (str): Премиум или нет вакансия
            employer_name (str): Название региона
            salary (str): Класс Зарплаты
        >>> Vacancy("Программист", "Компас-Плюс", "22.11.2022", "Описание", "Программирование", "between3And6", "Yes", "Новосибирск", Salary(10000, 20000.0, "Yes", 'RUR')).name
        ['Программист']
        >>> type(Vacancy("Программист", "Компас-Плюс", "22.11.2022", "Описание", "Программирование", "between3And6", "Yes", "Новосибирск", Salary(10000, 20000.0, "Yes", 'RUR'))).__name__
        'Vacancy'
        """
        self.name = [name]
        self.area_name = [area_name] 
        self.published_at = [published_at]
        self.description = [description]
        self.key_skills = key_skills
        self.experience_id = [experience_id]
        self.premium = [premium]
        self.employer_name = [employer_name]
        self.salary = salary
        self.elements=[name,description,key_skills,experience_id,premium,employer_name,salary,area_name,published_at]

class DataSet:
    """Класс который создает и рассчитывает значения для составления таблицы и статистики
    
    Attributes:
        report (class): Класс Report
        allDictionary (class): Класс AllDictionary
        file_name (str): Название файла со всеми профессиями
        profession_name (str): Название профессии, которую ищет работник
        vacancies_objects (list): Список со всеми профессиями
    """
    def __init__(self, profession = "None"):
        """Иницилизирует объект DataSet"""
        self.allDictionary = AllDictionary()
        self.profession_name = profession
        self.report = Report(self.profession_name)
        self.vacancies_objects = []

    def readystr(self, str_value):
        """Убирает в строке все лишние HTML теги
        
        Returns:
            str: Строка без HTML тегов
        
        >>> DataSet().readystr("<p>Hi</p>")
        'Hi'
        >>> DataSet().readystr("<html><head></head><p>Hi</p><body></body></html>")
        'Hi'
        """
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())
    
    def csv_ﬁlerAndReader(self, file_name):
        """Считывает файл, все ваканскии каждого года помещает в перенную investment, а все поля для этих ваканский в переменную names,
        также заполняяет вакансию с 9 полями(параметрами вакансии), вызывает функцию countsNumberAndSalary, которая считает данные для статистики
        
        Args:
            file_name (str): Название файла с профессиями по годам
        Returns:
            list: Возвращает список из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        names = []
        investment = []
        vacancies_objects = []
        with open("newCSV/" + file_name, encoding="UTF-8-sig") as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                if(len(names) == 0):
                    names = row
                elif(len(names) == len(row) and not("" in row)):
                    investment.append(row)

            for element in investment:
                salary = ["", "", "", ""]
                array = ["","","","","","","","",""]
                namesIndex = ["name", "area_name", "published_at", "description", "key_skills", "experience_id", "premium", "employer_name", "salary"]
                for item in range(len(names)):
                    value = element[item]
                    value = value.split("\n") if ("\n" in value) else [value]
                    arrayWithoutTegs = []
                    for word in value:
                        if(word.upper() == "TRUE"):
                            arrayWithoutTegs.append("Да")
                        elif(word.upper() == "FALSE"):
                            arrayWithoutTegs.append("Нет")
                        elif(word in fieldWorkExperience.keys()):
                            arrayWithoutTegs.append(fieldWorkExperience[word])
                        else:
                            arrayWithoutTegs.append(self.readystr(word))
                    if(len(arrayWithoutTegs) == 1):
                        arrayWithoutTegs = arrayWithoutTegs[0]
                    if(names[item] == "salary_from"):
                        salary[0] = arrayWithoutTegs
                    elif(names[item] == "salary_to"):
                        salary[1] = arrayWithoutTegs
                    elif(names[item] == "salary_gross"):
                        salary[2] = arrayWithoutTegs
                    elif(names[item] == "salary_currency"):
                        salary[3] = arrayWithoutTegs
                        readySalary = Salary(*salary)
                        array[8] = readySalary
                    else:
                        array[namesIndex.index(names[item])] = arrayWithoutTegs
                vacancy = Vacancy(*array)
                vacancies_objects.append(vacancy)
            listWithSumSalaryAndCount = self.countsNumberAndSalary({vacancy.published_at[0][0:4]:vacancies_objects})
            return listWithSumSalaryAndCount

    def countsNumberAndSalary(self, everyYear):
        """Функция считает количество ваканский в каждом году, сумму всех зарплат, количество ваканский(которую просят) в каждом году,
        сумму зарплат ваканский(которую просят) в каждом году
        
        Args:
            everyYear (dict): Словарь с ключом - год, а значение - список всех вакансий
        Returns:
            list: Возвращает список из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        sumAverageSalary = 0

        countAverageSalaryForProfession = 0
        sumAverageSalaryForProfession = 0
        for vacancie in list(everyYear.values())[0]:
            sumAverageSalary += int(currency_to_rub[vacancie.salary.salary_currency[0]] * (float(vacancie.salary.salary_from[0]) + float(vacancie.salary.salary_to[0])) / 2) 
            if(self.profession_name in vacancie.name[0]):
                countAverageSalaryForProfession += 1
                sumAverageSalaryForProfession += int(currency_to_rub[vacancie.salary.salary_currency[0]] * (float(vacancie.salary.salary_from[0]) + float(vacancie.salary.salary_to[0])) / 2) 
        return [list(everyYear.keys())[0],len(list(everyYear.values())[0]), sumAverageSalary, countAverageSalaryForProfession, sumAverageSalaryForProfession, list(everyYear.values())[0]]    
    
    def calculating_average_salary(self):
        """Функция считает значения среднии зарплаты в разных городах, долю ваканский в каждом городе относительно всех вакансий"""
        dicti3 = self.allDictionary.dictionary_level_salaries_cities
        for salaryYear in dicti3.keys():
            dicti3[salaryYear] = int(int(dicti3[salaryYear].split(" ")[1]) / int(dicti3[salaryYear].split(" ")[0]))
        dicti4 = self.allDictionary.dictionary_vacancy_rate_city
        for salaryYear in dicti4.keys(): 
            dicti4[salaryYear] = round(dicti4[salaryYear] / len(self.vacancies_objects), 4)

    def readyPrint(self, listWithSumSalaryAndCount):
        """Функция сначала считает заполняет словари с городами,
        потом выводит всю статистику, формирует графики,excel,pdf файлы
        
        Args:
            listWithSumSalaryAndCount (list): Cписок из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        for statisticForYear in listWithSumSalaryAndCount:
            self.vacancies_objects += statisticForYear[5]
            if(statisticForYear[0] in self.allDictionary.dictionary_salary_levels):
                #1 заполнения словаря Динамика уровня зарплат по годам
                self.allDictionary.dictionary_salary_levels[statisticForYear[0]] += int(statisticForYear[2] / statisticForYear[1])
                self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
                #2 заполнения словаря Динамика количества вакансий по годам
                self.allDictionary.dictionary_number_vacancies[statisticForYear[0]] += statisticForYear[1]
                self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
                #3 заполнения словаря Динамика уровня зарплат по годам для выбранной профессии
                self.allDictionary.dictionary_salary_levels_profession[statisticForYear[0]] += int(statisticForYear[4] / statisticForYear[3])
                self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
                #4 заполнения словаря Динамика количества вакансий по годам для выбранной профессии
                self.allDictionary.dictionary_number_vacancies_profession[statisticForYear[0]] += statisticForYear[3]
                self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession
            else:
                #1 заполнения словаря Динамика уровня зарплат по годам
                self.allDictionary.dictionary_salary_levels[statisticForYear[0]] = int(statisticForYear[2] / statisticForYear[1])
                self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
                #2 заполнения словаря Динамика количества вакансий по годам
                self.allDictionary.dictionary_number_vacancies[statisticForYear[0]] = statisticForYear[1]
                self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
                #3 заполнения словаря Динамика уровня зарплат по годам для выбранной профессии
                self.allDictionary.dictionary_salary_levels_profession[statisticForYear[0]] = int(statisticForYear[4] / statisticForYear[3])
                self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
                #4 заполнения словаря Динамика количества вакансий по годам для выбранной профессии
                self.allDictionary.dictionary_number_vacancies_profession[statisticForYear[0]] = statisticForYear[3]
                self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession

        for vacancie in self.vacancies_objects:
            #заполнения словаря с средней зарплатой по городам
            if (vacancie.area_name[0] in self.allDictionary.dictionary_level_salaries_cities.keys()):
                count = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[0]) + 1
                money = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[1]) + int(currency_to_rub[vacancie.salary.salary_currency[0]] * (float(vacancie.salary.salary_from[0]) + float(vacancie.salary.salary_to[0])) / 2) 
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = str(count) + " " + str(money)  
            else:
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = "1" + " " + str(int(currency_to_rub[vacancie.salary.salary_currency[0]] * (float(vacancie.salary.salary_from[0]) + float(vacancie.salary.salary_to[0])) / 2) )
            #заполнения словаря: доля ваканский каждого города относительно всех вакансий
            if (vacancie.area_name[0] in self.allDictionary.dictionary_vacancy_rate_city.keys()):
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] += 1
            else:
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] = 1
        self.calculating_average_salary()

        print("Динамика уровня зарплат по годам: ",end="")
        self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels = dict(sorted(self.allDictionary.dictionary_salary_levels.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_salary_levels)
        print("Динамика количества вакансий по годам: ",end="")
        self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies = dict(sorted(self.allDictionary.dictionary_number_vacancies.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_number_vacancies)
        print("Динамика уровня зарплат по годам для выбранной профессии: ",end="")
        self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession = dict(sorted(self.allDictionary.dictionary_salary_levels_profession.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_salary_levels_profession)
        print("Динамика количества вакансий по годам для выбранной профессии: ",end="")
        self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession = dict(sorted(self.allDictionary.dictionary_number_vacancies_profession.items(), key=lambda x: x[0]))
        print(self.allDictionary.dictionary_number_vacancies_profession)  
        print("Уровень зарплат по городам (в порядке убывания): ",end="")
        self.dictionary_level_salaries_cities = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_level_salaries_cities.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_level_salaries_cities = self.dictionary_level_salaries_cities
        print(self.dictionary_level_salaries_cities)
        print("Доля вакансий по городам (в порядке убывания): ",end="")
        self.dictionary_vacancy_rate_city = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_vacancy_rate_city.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_vacancy_rate_city = self.dictionary_vacancy_rate_city
        print(self.dictionary_vacancy_rate_city)
        
        self.report.generate_pdf()
        self.report.generate_graphics()
        self.report.generate_excel()

    def runningFunctionsInMultiThread(self, fileNames):
        """Функция запускает другие функции в многопотоке
        Args:
            fileNames (list): Список с названиями всех файлов csv
        Returns:
            listWithSumSalaryAndCount (list): Cписок из 6 элементов: Год, количество всех вакансий, сумму зарплат всех вакансий, количество нужных вакансий, сумму зарплат нужных вакансий, все вакансии
        """
        listWithSumSalaryAndCount = []
        with concurrent.futures.ProcessPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(self.csv_ﬁlerAndReader, fileName): fileName for fileName in fileNames}
            for fut in concurrent.futures.as_completed(futures):
                listWithSumSalaryAndCount.append(fut.result())
        return listWithSumSalaryAndCount

class InputConnect:
    """При помощи класса можно фильтровать, сортировать, выводить таблицу
    
    Attribute:
        data (class): Класс DataSet
    """
    def __init__(self):
        """Иницилизирует объект InputConnect"""
        self.data = DataSet()

    """Преобразование даты в нужный формат"""
    """def testDateTimeOne(self,published_at):
        return published_at[0].split("T")[0].split("-")[2] + "." + published_at[0].split("T")[0].split("-")[1] + "." + published_at[0].split("T")[0].split("-")[0]"""
    def testDateTimeTwo(self,published_at):
        date = published_at[0].replace("T","").replace("-","").replace(":","").replace("+","")[0:8] 
        return date[6:8] + "." + date[4:6] + "." + date[0:4]
    """def testDateTimeThree(self,published_at):
        date = datetime.strptime(published_at[0].replace("+",".").replace("T"," "), '%Y-%m-%d %H:%M:%S.%f')
        return ("0" + str(date.day))[-2:] + "." + str(date.month) + "." + str(date.year)
    def testDateTimeFour(self,published_at):
        date = str(datetime.date(datetime.strptime(published_at[0].replace("+",".").replace("T"," "), '%Y-%m-%d %H:%M:%S.%f')))
        return date[-2:] + "." + date[5:7] + "." + date[0:4]"""

    def filterVacancies(self, vacancy, filterVacansie):
        """Фильтрация вакансий в таблице
        
        Args:
            vacancy (class): Класс со всеми вакансиями
        Returns:
            list: возвращает вакансию, если она подходит, пустоту-если нет
            
        >>> filterVacansie = ["Идентификатор валюты оклада", "Рубли"]
        >>> vacancy = Vacancy("Программист", "Компас-Плюс", "22.11.2022", "Описание", "Программирование", "between3And6", "Yes", "Новосибирск", Salary(10000, 20000.0, "Yes", 'EUR'))
        >>> InputConnect().filterVacancies(vacancy, filterVacansie)
        {}
        """
        if(filterVacansie[0] != ""):
            article = fieldToEng[filterVacansie[0]]
            information = filterVacansie[1]
            if(article == "key_skills" and set(str(information).split(", ")).issubset(set(vacancy.key_skills))):
                return vacancy
            if(article == "currency" and "Идентификатор" not in filterVacansie[0] and int(float(vacancy.salary.salary_from[0])) <= int(information) <= int(float(vacancy.salary.salary_to[0]))):
                return vacancy
            if(article == "currency" and information == str(vacancy.salary.salary).split("(")[1].split(")")[0]):
                return vacancy
            if(article != "currency" and getattr(vacancy,article) == information):
                return vacancy
            if(article == "published_at" and self.testDateTimeTwo(vacancy.published_at) == information):
                return vacancy
            return {}
        return vacancy
    
    def sorti(self, data_vacancies):
        """Сортировка вакансий в таблице
        
        Args:
            data_vacancies (list): Список со всеми вакансиями
        Returns:
            list: возвращет список с отсортированными вакансиями
        """
        newData = data_vacancies
        if(sortingParameter == "Оклад"):
            newData = sorted(data_vacancies, key = lambda x: (currency_to_rub[x.salary.salary_currency] * (float(x.salary.salary_from) + float(x.salary.salary_to)) / 2),reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Дата публикации вакансии"):
            newData = sorted(data_vacancies, key = lambda x: x.published_at,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Навыки"):
            newData = sorted(data_vacancies, key = lambda x: len(x.key_skills) if type(x.key_skills) == list else 1, reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Опыт работы"):
            newData = sorted(data_vacancies, key = lambda x: x.experience_id[3],reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Премиум-вакансия"):   
            newData = sorted(data_vacancies, key = lambda x: "Да" if x.premium.upper()=="TRUE" else "Да",reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Название"):
            newData = sorted(data_vacancies, key = lambda x: x.name,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Описание"):
            newData = sorted(data_vacancies, key = lambda x: x.description,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Компания"):
            newData = sorted(data_vacancies, key = lambda x: x.employer_name,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter != ""):  
            newData = sorted(data_vacancies, key = lambda x: x.area_name,reverse = True if reverseOrder == "Да" else False)
        return newData
   
    def print_vacancies(self, data_vacancies, dic_naming):
        """Создание таблицы с вакансиями с возможностью сортировки,поиска по какому-то параметру
        
        Args:
            data_vacancies (str): Список со всеми вакансиями
            dic_naming (str): Список со всеми параметрами
        """
        table = PrettyTable()   
        table.field_names = ["№"] + list(dic_naming.values())
        table._max_width = {"Название":20, "Описание":20, "Навыки":20, "Опыт работы":20, "Премиум-вакансия":20, "Компания":20, "Оклад":20, "Название региона":20, "Дата публикации вакансии":20}
        table.hrules = 1
        table.align = "l"
        number = 0
        data_vacancies = self.sorti(data_vacancies)
        for vacancy in data_vacancies:
            vacancy = self.filterVacancies(vacancy, filterVacansie)
            number += 1
            val = [str(number)]
            if(type(vacancy) != dict):
                for item in vacancy.elements:
                    if(type(item) == list):
                        element = ('\n'.join(str(x) for x in item))
                    elif(type(item) == Salary):
                        element = item.salary
                    else:
                        element = item    
                    val.append(str(element) if (len(element) < 100) else (element[0:100] + "..."))
                date = val.pop().split("T")[0]
                val.append(date.split("-")[2] + "." + date.split("-")[1] + "." + date.split("-")[0])
            if(len(val) > 1):
                table.add_row(val)
            else:
                number -= 1
        if(number == 0):
            print("Ничего не найдено")
        else:
            print(table.get_string(fields = (["№"] + (column if len(column) > 1 else list(dic_naming.values()))), start = (int(lines[0]) - 1 if len(lines) > 0 else 0), end = (int(lines[1]) - 1 if len(lines) > 1 else number)))

    def readyPrint(self):
        """Проверяет введены и корретны ли обязательные поля"""
        if(os.stat(name).st_size == 0):
            print("Пустой файл")
        elif(len(filterVacansie[0]) > 0 and len(filterVacansie) == 1):
            print("Формат ввода некорректен")
        elif(filterVacansie[0] not in fieldToEng.keys() and filterVacansie[0] != ""):
            print("Параметр поиска некорректен")
        elif(sortingParameter not in fieldToEng.keys() and sortingParameter != ""):
            print("Параметр сортировки некорректен")
        elif(reverseOrder != "Да" and reverseOrder != "Нет" and reverseOrder != ""):
            print("Порядок сортировки задан некорректно")
        else:
            a, b = self.data.сsv_reader(name)
            if len(b) == 0 or len(a) == 0:
                print("Нет данных")
            else:
                self.print_vacancies(self.data.csv_ﬁler(a, b), fieldToRus)  

class Report:
    """Класс который создает и рассчитывает значения для составления таблицы и статистики
    
    Attributes:
        dictionary_salary_levels (dict): Динамика уровня зарплат по годам
        dictionary_number_vacancies (dict): Динамика количества вакансий по годам
        dictionary_salary_levels_profession (dict): Динамика уровня зарплат по годам для выбранной профессии
        dictionary_number_vacancies_profession (dict): Динамика количества вакансий по годам для выбранной профессии
        dictionary_level_salaries_cities (dict): Уровень зарплат по городам (в порядке убывания)
        dictionary_vacancy_rate_city (dict): Доля вакансий по городам (в порядке убывания)
    """
    def __init__(self, profession):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        self.profession = profession

    def generate_graphics(self):
        """Генерация всех графиков в картинку при помоще библиотеки matplotlib"""
        plt.rcParams.update({'font.size': 8})
        #1 ГРАФИК
        labelsLevel = self.dictionary_salary_levels.keys()
        middle_salary = self.dictionary_salary_levels.values()
        middle_salary_profession = self.dictionary_salary_levels_profession.values()

        x = np.arange(len(labelsLevel))  
        width = 0.35 

        ax = plt.subplot(221)
        ax.bar(x - width/2, middle_salary, width, label='средняя з/п')
        ax.bar(x + width/2, middle_salary_profession, width, label='з/п ' + self.profession)
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(x, labelsLevel, rotation = 90)
        plt.grid(axis='y')
        ax.legend()
        #2 ГРАФИК
        labelsCount = self.dictionary_number_vacancies.keys()
        middle_count = self.dictionary_number_vacancies.values()
        middle_count_profession = self.dictionary_number_vacancies_profession.values()

        x = np.arange(len(labelsCount))  
        width = 0.35 

        ax = plt.subplot(222)
        ax.bar(x - width/2, middle_count, width, label='Количества вакансий')
        ax.bar(x + width/2, middle_count_profession, width, label='Количества вакансий \n' + self.profession)
        ax.set_title('Количество вакансий по годам')
        ax.set_xticks(x, labelsCount, rotation = 90)
        plt.grid(axis='y')
        ax.legend()
        #3 ГРАФИК
        ax = plt.subplot(223)
        ax.set_title('Уровень зарплат по городам')
        plt.rcParams.update({'font.size': 6})
        city = self.dictionary_level_salaries_cities.keys()
        y_pos = np.arange(len(city))
        performance = self.dictionary_level_salaries_cities.values()

        ax.barh(y_pos, performance,  align = 'center')
        ax.set_yticks(y_pos, labels = city)
        ax.invert_yaxis() 
        plt.grid(axis='x')

        #4 ГРАФИК   
        ax = plt.subplot(224)
        plt.rcParams.update({'font.size': 8})
        ax.set_title('Доля вакансий по городам')
        plt.rcParams.update({'font.size': 6})
        ax.pie(list(self.dictionary_vacancy_rate_city.values()) + [1 - sum(self.dictionary_vacancy_rate_city.values())], labels = list(self.dictionary_vacancy_rate_city.keys()) + ["Другие"])

        plt.subplots_adjust(wspace = 0.5, hspace = 0.5)
        plt.savefig("Graphics", dpi = 200, bbox_inches='tight')

    def set_border(self, ws, cell_range):
        """Генерация всех графиков в картинку
        
        Args:
            ws (worksheet): рабочая страница в Excel 
            cell_range (str): строка, в которой переданы ячейки, которые нудно изменить 
        """
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)    

    def set_bold(self, ws, cell_range):
        """Делает жирный шрифт в шапке таблицы
        Args:
            ws (worksheet): рабочая страница в Excel 
            cell_range (str): строка, в которой переданы ячейки, которые нудно изменить    
        """
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold = True)

    def width(self, ws):
        """Задаёт размеры каждого столбца в таблице
        
        Args:
            ws (worksheet): рабочая страница в Excel 
        """
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.5
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate_excel(self):
        """Генерирует таблицу Excel"""
        wb = Workbook()
        #1 СТРАНИЦА
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws1.append(["Год", "Средняя зарплата", "Средняя зарплата - " + self.profession, "Количество вакансий", "Количество вакансий - " + self.profession])
        
        countIndexA = 2
        for year in self.dictionary_salary_levels.keys():
            ws1['A' + str(countIndexA)] = year
            countIndexA += 1
        countIndexB = 2
        for value in self.dictionary_salary_levels.values():
            ws1['B' + str(countIndexB)] = value
            countIndexB += 1
        countIndexC = 2
        for value in self.dictionary_salary_levels_profession.values():
            ws1['C' + str(countIndexC)] = value
            countIndexC += 1
        countIndexD = 2
        for value in self.dictionary_number_vacancies.values():
            ws1['D' + str(countIndexD)] = value
            countIndexD += 1
        countIndexE = 2
        for value in self.dictionary_number_vacancies_profession.values():
            ws1['E' + str(countIndexE)] = value
            countIndexE += 1
        
        self.width(ws1)
        self.set_border(ws1, 'A1:E' + str(1 + len(self.dictionary_salary_levels.keys()))) 
        self.set_bold(ws1, 'A1:E1') 
        #2 СТРАНИЦА
        ws2 = wb.create_sheet("Статистика по городам")
        ws2.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        
        countWS2IndexA = 2
        for city in self.dictionary_level_salaries_cities.keys():
            ws2['A' + str(countWS2IndexA)] = city
            countWS2IndexA += 1
        countWS2IndexB = 2
        for value in self.dictionary_level_salaries_cities.values():
            ws2['B' + str(countWS2IndexB)] = value
            countWS2IndexB += 1
        countWS2IndexD = 2
        for city in self.dictionary_vacancy_rate_city.keys():
            ws2['D' + str(countWS2IndexD)] = city
            countWS2IndexD += 1
        countWS2IndexE = 2
        for value in self.dictionary_vacancy_rate_city.values():
            ws2['E' + str(countWS2IndexE)] = value
            ws2['E' + str(countWS2IndexE)].number_format = '0.00%'
            countWS2IndexE += 1

        self.set_border(ws2, 'A1:B' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_border(ws2, 'D1:E' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_bold(ws2, 'A1:E1') 
        self.width(ws2)
        wb.save('report.xlsx')

    def generate_pdf(self):
        """Генерирует файл PDF"""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")
        
        pdf_template = template.render({"name": self.profession})
        config = pdfkit.configuration(wkhtmltopdf = r'D:\Учеба\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf_template = pdf_template.replace("path", os.path.abspath(os.curdir))
        options = {'enable-local-file-access': None}

        readyHtmlTable = self.generate_table()
        pdf_template = pdf_template.replace("$table1;", readyHtmlTable)
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def generate_table(self):
        """Создаёт таблицу для HTML кода
        
        >>> profession = "Программист"
        >>> Report(profession).generate_table()
        "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - Программист</th><th>Количество вакансий</th><th>Количество вакансий - Программист</th></tr></tr></table><h1>Статистика по городам</h1><table class='tableCity1'><tr><th>Город</th><th>Уровень зарплат</th></tr></table><table class='tableCity2'><tr><th>Город</th><th>Уровень зарплат</th></tr></table>"
        """
        #1 ТАБЛИЦА
        readyHtmlTable = "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - "
        readyHtmlTable += self.profession + "</th><th>Количество вакансий</th><th>Количество вакансий - " + self.profession + "</th></tr>"
        for index in range(len(list(self.dictionary_salary_levels.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels_profession.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies_profession.values())[index]) + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table><h1>Статистика по городам</h1>"
        #2 ТАБЛИЦА
        readyHtmlTable += "<table class='tableCity1'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.dictionary_level_salaries_cities.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_level_salaries_cities.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_level_salaries_cities.values())[index]) + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table>"
        #3 ТАБЛИЦА
        readyHtmlTable += "<table class='tableCity2'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.dictionary_vacancy_rate_city.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_vacancy_rate_city.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(round(list(self.dictionary_vacancy_rate_city.values())[index] * 100, 2)) + "%" + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table>"
        return readyHtmlTable
def main():
    choice = input("Введите Вакансии/Статистика: ")
   
    if (choice == "Вакансии"):
        name = input("Введите название файла: ")
        filterVacansie = input("Введите параметр фильтрации: ").split(": ")
        sortingParameter = input("Введите параметр сортировки: ")
        reverseOrder = input("Обратный порядок сортировки (Да / Нет): ")
        lines = input("Введите диапазон вывода: ").split()
        column = input("Введите требуемые столбцы: ").split(", ")
        
    """Итоговый вывод в зависимости от выбора"""
    if(choice == "Статистика"):
        folderName = input("Введите название папки: ")
        profession = input("Введите название профессии: ")
        fileNames = [] 

        for (dirpath, dirnames, filenames) in walk(folderName):
            fileNames.extend(filenames)
            break

        printer = DataSet(profession)
        listWithSumSalaryAndCount = printer.runningFunctionsInMultiThread(fileNames)   
        printer.readyPrint(listWithSumSalaryAndCount)
    if(choice == "Вакансии"):
        printer = InputConnect()
        printer.readyPrint()
if __name__ == "__main__":
    doctest.testmod()
    cProfile.run('main()')