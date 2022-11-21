import csv
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font

name = input("Введите название файла: ") 
profession = input("Введите название профессии: ")

names = []
investment = []

currency_to_rub = {  
    "AZN": 35.68,  
    "BYR": 23.91,  
    "EUR": 59.90,  
    "GEL": 21.74,  
    "KGS": 0.76,  
    "KZT": 0.13,  
    "RUR": 1,  
    "UAH": 1.64,  
    "USD": 60.66,  
    "UZS": 0.0055,  
}

class AllDictionary:
    def __init__(self):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        
class Vacancy:
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at, salary_gross, description, key_skills, experience_id, premium, employer_name, salary):
        self.name = name
        self.salary_from = salary_from 
        self.salary_to = salary_to
        self.salary_currency = salary_currency 
        self.mysalary = int(currency_to_rub[salary_currency[0]] * (float(salary_from[0]) + float(salary_to[0])))
        self.area_name = area_name 
        self.published_at = published_at
        self.salary_gross = salary_gross
        self.published_at = published_at
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        
class DataSet:
    def __init__(self):
        self.report = Report()
        self.allDictionary = AllDictionary()
        self.file_name = name
        self.profession_name = profession
        self.vacancies_objects = []

    def сsv_reader(self,ﬁle_name): 
        global names
        with open(ﬁle_name, encoding="UTF-8-sig") as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                if(len(names) == 0):
                    names = row
                elif(len(names) == len(row) and not("" in row)):
                    investment.append(row)
        return investment, names

    def csv_ﬁler(self, reader):
        for element in reader:
            array = ["","","","","","","","","","","","",""]
            namesIndex = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at", "salary_gross", "description", "key_skills", "experience_id", "premium", "employer_name", "salary"]
            for item in range(len(names)):
                value = element[item]
                value = value.split("\n") if ("\n" in value) else [value]
                array[namesIndex.index(names[item])] = value
            vacancy = Vacancy(*array)
            self.vacancies_objects.append(vacancy)
        return self.vacancies_objects

    def completion_dictionary(self):
        for vacancie in self.vacancies_objects:
            #1 вывод
            if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_salary_levels.keys()):
                count = int(self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])].split(" ")[0]) + 1
                money = int(self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])].split(" ")[1]) + vacancie.mysalary
                self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])] = str(count) + " " + str(money)
            else:
                self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = "0" + " " + "0"
                self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])] = "1" + " " + str(vacancie.mysalary)
            #2 вывод
            if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_number_vacancies.keys()):
                self.allDictionary.dictionary_number_vacancies[int(vacancie.published_at[0][0:4])] += 1
            else:
                self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] = 0
                self.allDictionary.dictionary_number_vacancies[int(vacancie.published_at[0][0:4])] = 1
            #3 вывод
            if(profession in vacancie.name[0]):
                if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_salary_levels_profession.keys()):
                    count = int(self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])].split(" ")[0]) + 1
                    money = int(self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])].split(" ")[1]) + vacancie.mysalary
                    self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = str(count) + " " + str(money)
                else:
                    self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = "1" + " " + str(vacancie.mysalary)
            #4 вывод
            if(profession in vacancie.name[0]):
                if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_number_vacancies_profession.keys()):
                    self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] += 1
                else:
                    self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] = 1
            #5 вывод
            if (vacancie.area_name[0] in self.allDictionary.dictionary_level_salaries_cities.keys()):
                count = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[0]) + 1
                money = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[1]) + vacancie.mysalary
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = str(count) + " " + str(money)  
            else:
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = "1" + " " + str(vacancie.mysalary)
            #6 вывод
            if (vacancie.area_name[0] in self.allDictionary.dictionary_vacancy_rate_city.keys()):
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] += 1
            else:
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] = 1

    def calculating_average_salary(self):
        dicti = self.allDictionary.dictionary_salary_levels
        for salaryYear in dicti.keys():
            dicti[salaryYear] = int(int(dicti[salaryYear].split(" ")[1]) / (int(dicti[salaryYear].split(" ")[0]) * 2))
        dicti2 = self.allDictionary.dictionary_salary_levels_profession
        for salaryYear in dicti2.keys():
            if(int(dicti2[salaryYear].split(" ")[0]) != 0):
                dicti2[salaryYear] = int(int(dicti2[salaryYear].split(" ")[1]) / (int(dicti2[salaryYear].split(" ")[0]) * 2))
            else:
                dicti2[salaryYear] = 0
        dicti3 = self.allDictionary.dictionary_level_salaries_cities
        for salaryYear in dicti3.keys():
            dicti3[salaryYear] = int(int(dicti3[salaryYear].split(" ")[1]) / (int(dicti3[salaryYear].split(" ")[0]) * 2))
        dicti4 = self.allDictionary.dictionary_vacancy_rate_city
        for salaryYear in dicti4.keys():
            dicti4[salaryYear] = round(dicti4[salaryYear] / len(self.vacancies_objects), 4)

    def readyPrint(self):
        a, b = self.сsv_reader(name)
        self.csv_ﬁler(a)
        self.completion_dictionary()
        self.calculating_average_salary()
        print("Динамика уровня зарплат по годам: ",end="")
        self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
        print(self.allDictionary.dictionary_salary_levels)
        print("Динамика количества вакансий по годам: ",end="")
        self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
        print(self.allDictionary.dictionary_number_vacancies)
        print("Динамика уровня зарплат по годам для выбранной профессии: ",end="")
        self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
        print(self.allDictionary.dictionary_salary_levels_profession)
        print("Динамика количества вакансий по годам для выбранной профессии: ",end="")
        self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession
        print(self.allDictionary.dictionary_number_vacancies_profession)   
        print("Уровень зарплат по городам (в порядке убывания): ",end="")
        self.dictionary_level_salaries_cities = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_level_salaries_cities.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_level_salaries_cities = self.dictionary_level_salaries_cities
        print(self.dictionary_level_salaries_cities)
        print("Доля вакансий по городам (в порядке убывания): ",end="")
        self.dictionary_vacancy_rate_city = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_vacancy_rate_city.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_vacancy_rate_city = self.dictionary_vacancy_rate_city
        print(self.dictionary_vacancy_rate_city)
        self.report.generate_excel()



class Report:
    def __init__(self):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}

    def set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)    

    def set_bold(self, ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold = True)

    def width(self, ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.5
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate_excel(self):
        wb = Workbook()
        #1 СТРАНИЦА
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws1.append(["Год", "Средняя зарплата", "Средняя зарплата - " + profession, "Количество вакансий", "Количество вакансий - " + profession])
        
        countIndexA = 2
        for year in self.dictionary_salary_levels.keys():
            ws1['A' + str(countIndexA)] = year
            countIndexA += 1
        countIndexB = 2
        for value in self.dictionary_salary_levels.values():
            ws1['B' + str(countIndexB)] = value
            countIndexB += 1
        countIndexC = 2
        for value in self.dictionary_salary_levels_profession.values():
            ws1['C' + str(countIndexC)] = value
            countIndexC += 1
        countIndexD = 2
        for value in self.dictionary_number_vacancies.values():
            ws1['D' + str(countIndexD)] = value
            countIndexD += 1
        countIndexE = 2
        for value in self.dictionary_number_vacancies_profession.values():
            ws1['E' + str(countIndexE)] = value
            countIndexE += 1
        
        self.width(ws1)
        self.set_border(ws1, 'A1:E' + str(1 + len(self.dictionary_salary_levels.keys()))) 
        self.set_bold(ws1, 'A1:E1') 
        #2 СТРАНИЦА
        ws2 = wb.create_sheet("Статистика по городам")
        ws2.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        
        countWS2IndexA = 2
        for city in self.dictionary_level_salaries_cities.keys():
            ws2['A' + str(countWS2IndexA)] = city
            countWS2IndexA += 1
        countWS2IndexB = 2
        for value in self.dictionary_level_salaries_cities.values():
            ws2['B' + str(countWS2IndexB)] = value
            countWS2IndexB += 1
        countWS2IndexD = 2
        for city in self.dictionary_vacancy_rate_city.keys():
            ws2['D' + str(countWS2IndexD)] = city
            countWS2IndexD += 1
        countWS2IndexE = 2
        for value in self.dictionary_vacancy_rate_city.values():
            ws2['E' + str(countWS2IndexE)] = value
            ws2['E' + str(countWS2IndexE)].number_format = '0.00%'
            countWS2IndexE += 1

        self.set_border(ws2, 'A1:B' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_border(ws2, 'D1:E' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_bold(ws2, 'A1:E1') 
        self.width(ws2)
        wb.save('report.xlsx')


printer = DataSet()
printer.readyPrint()
