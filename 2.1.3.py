import csv
import re
import os
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from jinja2 import Environment, FileSystemLoader
from prettytable import PrettyTable

<<<<<<< HEAD
choice = input("Enter Vacancies/Statistics: ")
=======
choice = input("Inserisci Lavori / Statistiche: ")
>>>>>>> 13f5a11cd16f794ef733872c146b7b51bbb2d49d
name = input("Введите название файла: ") 
profession = input("Введите название профессии: ")

if (choice == "Вакансии"):
    filterVacansie = input("Введите параметр фильтрации: ").split(": ")
    sortingParameter = input("Введите параметр сортировки: ")
    reverseOrder = input("Обратный порядок сортировки (Да / Нет): ") 
    lines = input("Введите диапазон вывода: ").split()
    column = input("Введите требуемые столбцы: ").split(", ")

names = []
investment = []

currency_to_rub = {  
    "AZN": 35.68,  
    "BYR": 23.91,  
    "EUR": 59.90,  
    "GEL": 21.74,  
    "KGS": 0.76,  
    "KZT": 0.13,  
    "RUR": 1,  
    "UAH": 1.64,  
    "USD": 60.66,  
    "UZS": 0.0055,  
}

fieldToEng = {
"Название": "name",
"Описание":"description",
"Навыки":"key_skills",
"Опыт работы":"experience_id",
"Премиум-вакансия":"premium",
"Компания":"employer_name",
"Оклад": "currency",
"Название региона":"area_name",
"Дата публикации вакансии":"published_at",
"Идентификатор валюты оклада":"currency"
}

fieldToRus = {
"name":"Название",
"description":"Описание",
"key_skills":"Навыки",
"experience_id":"Опыт работы",
"premium":"Премиум-вакансия",
"employer_name":"Компания",
"currency":"Оклад",
"area_name":"Название региона",
"published_at":"Дата публикации вакансии"
}

fieldCurrency = {
"AZN": "Манаты",
"BYR": "Белорусские рубли",
"EUR": "Евро",
"GEL": "Грузинский лари",
"KGS": "Киргизский сом",
"KZT": "Тенге",
"RUR": "Рубли",
"UAH": "Гривны",
"USD": "Доллары",
"UZS": "Узбекский сум"
}

fieldWorkExperience = {
"noExperience": "Нет опыта",
"between1And3": "От 1 года до 3 лет",
"between3And6": "От 3 до 6 лет",
"moreThan6": "Более 6 лет"
}

class AllDictionary:
    def __init__(self):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}
        
class Salary:
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        self.salary_from = [salary_from]
        self.salary_to = [salary_to]
        self.salary_gross = [salary_gross]
        self.salary_currency = [salary_currency]
        self.salary = str('{:,}'.format(int(float(salary_from))).replace(',', ' ')) + " - " + str('{:,}'.format(int(float(salary_to))).replace(',', ' ')) + " (" + fieldCurrency[salary_currency] + ") (" + ("Без вычета налогов" if salary_gross.upper() == "ДА" else "С вычетом налогов") + ")"

class Vacancy:
    def __init__(self, name, area_name, published_at, description, key_skills, experience_id, premium, employer_name, salary):
        self.name = [name]
        self.mysalary = int(currency_to_rub[salary.salary_currency[0]] * (float(salary.salary_from[0]) + float(salary.salary_to[0])))
        self.area_name = [area_name] 
        self.published_at = [published_at]
        self.published_at = [published_at]
        self.description = [description]
        self.key_skills = key_skills
        self.experience_id = [experience_id]
        self.premium = [premium]
        self.employer_name = [employer_name]
        self.salary = salary
        self.elements=[name,description,key_skills,experience_id,premium,employer_name,salary,area_name,published_at]

class DataSet:
    def __init__(self):
        self.report = Report()
        self.allDictionary = AllDictionary()
        self.file_name = name
        self.profession_name = profession
        self.vacancies_objects = []

    def readystr(str_value):
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())
    
    def сsv_reader(self,ﬁle_name): 
        global names
        with open(ﬁle_name, encoding="UTF-8-sig") as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                if(len(names) == 0):
                    names = row
                elif(len(names) == len(row) and not("" in row)):
                    investment.append(row)
        return investment, names

    def csv_ﬁler(self, reader, list_naming):
        for element in reader:
            salary = ["", "", "", ""]
            array = ["","","","","","","","",""]
            namesIndex = ["name", "area_name", "published_at", "description", "key_skills", "experience_id", "premium", "employer_name", "salary"]
            for item in range(len(names)):
                value = element[item]
                value = value.split("\n") if ("\n" in value) else [value]
                arrayWithoutTegs = []
                for word in value:
                    if(word.upper() == "TRUE"):
                        arrayWithoutTegs.append("Да")
                    elif(word.upper() == "FALSE"):
                        arrayWithoutTegs.append("Нет")
                    elif(word in fieldWorkExperience.keys()):
                        arrayWithoutTegs.append(fieldWorkExperience[word])
                    else:
                        arrayWithoutTegs.append(DataSet.readystr(word))
                if(len(arrayWithoutTegs) == 1):
                    arrayWithoutTegs = arrayWithoutTegs[0]
                if(list_naming[item] == "salary_from"):
                    salary[0] = arrayWithoutTegs
                elif(list_naming[item] == "salary_to"):
                    salary[1] = arrayWithoutTegs
                elif(list_naming[item] == "salary_gross"):
                    salary[2] = arrayWithoutTegs
                elif(list_naming[item] == "salary_currency"):
                    salary[3] = arrayWithoutTegs
                    readySalary = Salary(*salary)
                    array[8] = readySalary
                else:
                    array[namesIndex.index(list_naming[item])] = arrayWithoutTegs
            vacancy = Vacancy(*array)
            self.vacancies_objects.append(vacancy)
        return self.vacancies_objects

    def print_vacancies(self, data_vacancies, dic_naming):
        table = PrettyTable()   
        table.field_names = ["№"] + list(dic_naming.values())
        table._max_width = {"Название":20, "Описание":20, "Навыки":20, "Опыт работы":20, "Премиум-вакансия":20, "Компания":20, "Оклад":20, "Название региона":20, "Дата публикации вакансии":20}
        table.hrules = 1
        table.align = "l"
        number = 0
        data_vacancies = self.sorti(data_vacancies)
        for vacancy in data_vacancies:
            vacancy = self.filterVacancies(vacancy)
            number += 1
            val = [str(number)]
            if(type(vacancy) != dict):
                for item in vacancy.elements:
                    if(type(item) == list):
                        element = ('\n'.join(str(x) for x in item))
                    elif(type(item) == Salary):
                        element = item.salary
                    else:
                        element = item    
                    val.append(str(element) if (len(element) < 100) else (element[0:100] + "..."))
                date = val.pop().split("T")[0]
                val.append(date.split("-")[2] + "." + date.split("-")[1] + "." + date.split("-")[0])
            if(len(val) > 1):
                table.add_row(val)
            else:
                number -= 1
        if(number == 0):
            print("Ничего не найдено")
        else:
            print(table.get_string(fields = (["№"] + (column if len(column) > 1 else list(dic_naming.values()))), start = (int(lines[0]) - 1 if len(lines) > 0 else 0), end = (int(lines[1]) - 1 if len(lines) > 1 else number)))

    def completion_dictionary(self):
        for vacancie in self.vacancies_objects:
            #1 вывод
            if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_salary_levels.keys()):
                count = int(self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])].split(" ")[0]) + 1
                money = int(self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])].split(" ")[1]) + vacancie.mysalary
                self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])] = str(count) + " " + str(money)
            else:
                self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = "0" + " " + "0"
                self.allDictionary.dictionary_salary_levels[int(vacancie.published_at[0][0:4])] = "1" + " " + str(vacancie.mysalary)
            #2 вывод
            if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_number_vacancies.keys()):
                self.allDictionary.dictionary_number_vacancies[int(vacancie.published_at[0][0:4])] += 1
            else:
                self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] = 0
                self.allDictionary.dictionary_number_vacancies[int(vacancie.published_at[0][0:4])] = 1
            #3 вывод
            if(profession in vacancie.name[0]):
                if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_salary_levels_profession.keys()):
                    count = int(self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])].split(" ")[0]) + 1
                    money = int(self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])].split(" ")[1]) + vacancie.mysalary
                    self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = str(count) + " " + str(money)
                else:
                    self.allDictionary.dictionary_salary_levels_profession[int(vacancie.published_at[0][0:4])] = "1" + " " + str(vacancie.mysalary)
            #4 вывод
            if(profession in vacancie.name[0]):
                if (int(vacancie.published_at[0][0:4]) in self.allDictionary.dictionary_number_vacancies_profession.keys()):
                    self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] += 1
                else:
                    self.allDictionary.dictionary_number_vacancies_profession[int(vacancie.published_at[0][0:4])] = 1
            #5 вывод
            if (vacancie.area_name[0] in self.allDictionary.dictionary_level_salaries_cities.keys()):
                count = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[0]) + 1
                money = int(self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]].split(" ")[1]) + vacancie.mysalary
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = str(count) + " " + str(money)  
            else:
                self.allDictionary.dictionary_level_salaries_cities[vacancie.area_name[0]] = "1" + " " + str(vacancie.mysalary)
            #6 вывод
            if (vacancie.area_name[0] in self.allDictionary.dictionary_vacancy_rate_city.keys()):
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] += 1
            else:
                self.allDictionary.dictionary_vacancy_rate_city[vacancie.area_name[0]] = 1

    def calculating_average_salary(self):
        dicti = self.allDictionary.dictionary_salary_levels
        for salaryYear in dicti.keys():
            dicti[salaryYear] = int(int(dicti[salaryYear].split(" ")[1]) / (int(dicti[salaryYear].split(" ")[0]) * 2))
        dicti2 = self.allDictionary.dictionary_salary_levels_profession
        for salaryYear in dicti2.keys():
            if(int(dicti2[salaryYear].split(" ")[0]) != 0):
                dicti2[salaryYear] = int(int(dicti2[salaryYear].split(" ")[1]) / (int(dicti2[salaryYear].split(" ")[0]) * 2))
            else:
                dicti2[salaryYear] = 0
        dicti3 = self.allDictionary.dictionary_level_salaries_cities
        for salaryYear in dicti3.keys():
            dicti3[salaryYear] = int(int(dicti3[salaryYear].split(" ")[1]) / (int(dicti3[salaryYear].split(" ")[0]) * 2))
        dicti4 = self.allDictionary.dictionary_vacancy_rate_city
        for salaryYear in dicti4.keys():
            dicti4[salaryYear] = round(dicti4[salaryYear] / len(self.vacancies_objects), 4)

    def readyPrint(self):
        a, b = self.сsv_reader(name)

        self.csv_ﬁler(a, b)
        self.completion_dictionary()
        self.calculating_average_salary()
        print("Динамика уровня зарплат по годам: ",end="")
        self.report.dictionary_salary_levels = self.allDictionary.dictionary_salary_levels
        print(self.allDictionary.dictionary_salary_levels)
        print("Динамика количества вакансий по годам: ",end="")
        self.report.dictionary_number_vacancies = self.allDictionary.dictionary_number_vacancies
        print(self.allDictionary.dictionary_number_vacancies)
        print("Динамика уровня зарплат по годам для выбранной профессии: ",end="")
        self.report.dictionary_salary_levels_profession = self.allDictionary.dictionary_salary_levels_profession
        print(self.allDictionary.dictionary_salary_levels_profession)
        print("Динамика количества вакансий по годам для выбранной профессии: ",end="")
        self.report.dictionary_number_vacancies_profession = self.allDictionary.dictionary_number_vacancies_profession
        print(self.allDictionary.dictionary_number_vacancies_profession)   
        print("Уровень зарплат по городам (в порядке убывания): ",end="")
        self.dictionary_level_salaries_cities = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_level_salaries_cities.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_level_salaries_cities = self.dictionary_level_salaries_cities
        print(self.dictionary_level_salaries_cities)
        print("Доля вакансий по городам (в порядке убывания): ",end="")
        self.dictionary_vacancy_rate_city = dict(sorted(filter(lambda x: self.allDictionary.dictionary_vacancy_rate_city[x[0]] >= 0.01, self.allDictionary.dictionary_vacancy_rate_city.items()), key=lambda item: item[1], reverse=True)[:10])
        self.report.dictionary_vacancy_rate_city = self.dictionary_vacancy_rate_city
        print(self.dictionary_vacancy_rate_city)
        self.report.generate_pdf()
        self.report.generate_graphics()
        self.report.generate_excel()


class InputConnect:
    def __init__(self):
        self.data = DataSet()

    def filterVacancies(self, vacancy):
        if(filterVacansie[0] != ""):
            article = fieldToEng[filterVacansie[0]]
            information = filterVacansie[1]
            if(article == "key_skills" and set(str(information).split(", ")).issubset(set(vacancy.key_skills))):
                return vacancy
            if(article == "currency" and "Идентификатор" not in filterVacansie[0] and int(float(vacancy.salary.salary_from)) <= int(information) <= int(float(vacancy.salary.salary_to))):
                return vacancy
            if(article == "currency" and information == str(vacancy.salary.salary).split("(")[1].split(")")[0]):
                return vacancy
            if(article != "currency" and getattr(vacancy,article) == information):
                return vacancy
            if(article == "published_at" and vacancy.published_at.split("T")[0].split("-")[2] + "." + vacancy.published_at.split("T")[0].split("-")[1] + "." + vacancy.published_at.split("T")[0].split("-")[0] == information):
                return vacancy
            return {}
        return vacancy
    
    def sorti(self, data_vacancies):
        newData = data_vacancies
        if(sortingParameter == "Оклад"):
            newData = sorted(data_vacancies, key = lambda x: (currency_to_rub[x.salary.salary_currency] * (float(x.salary.salary_from) + float(x.salary.salary_to)) / 2),reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Дата публикации вакансии"):
            newData = sorted(data_vacancies, key = lambda x: x.published_at,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Навыки"):
            newData = sorted(data_vacancies, key = lambda x: len(x.key_skills) if type(x.key_skills) == list else 1, reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Опыт работы"):
            newData = sorted(data_vacancies, key = lambda x: x.experience_id[3],reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Премиум-вакансия"):   
            newData = sorted(data_vacancies, key = lambda x: "Да" if x.premium.upper()=="TRUE" else "Да",reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Название"):
            newData = sorted(data_vacancies, key = lambda x: x.name,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Описание"):
            newData = sorted(data_vacancies, key = lambda x: x.description,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter == "Компания"):
            newData = sorted(data_vacancies, key = lambda x: x.employer_name,reverse = True if reverseOrder == "Да" else False)
        elif(sortingParameter != ""):  
            newData = sorted(data_vacancies, key = lambda x: x.area_name,reverse = True if reverseOrder == "Да" else False)
        return newData
   
    def print_vacancies(self, data_vacancies, dic_naming):
        table = PrettyTable()   
        table.field_names = ["№"] + list(dic_naming.values())
        table._max_width = {"Название":20, "Описание":20, "Навыки":20, "Опыт работы":20, "Премиум-вакансия":20, "Компания":20, "Оклад":20, "Название региона":20, "Дата публикации вакансии":20}
        table.hrules = 1
        table.align = "l"
        number = 0
        data_vacancies = self.sorti(data_vacancies)
        for vacancy in data_vacancies:
            vacancy = self.filterVacancies(vacancy)
            number += 1
            val = [str(number)]
            if(type(vacancy) != dict):
                for item in vacancy.elements:
                    if(type(item) == list):
                        element = ('\n'.join(str(x) for x in item))
                    elif(type(item) == Salary):
                        element = item.salary
                    else:
                        element = item    
                    val.append(str(element) if (len(element) < 100) else (element[0:100] + "..."))
                date = val.pop().split("T")[0]
                val.append(date.split("-")[2] + "." + date.split("-")[1] + "." + date.split("-")[0])
            if(len(val) > 1):
                table.add_row(val)
            else:
                number -= 1
            print(vacancy.name)
        if(number == 0):
            print("Ничего не найдено")
        else:
            print(table.get_string(fields = (["№"] + (column if len(column) > 1 else list(dic_naming.values()))), start = (int(lines[0]) - 1 if len(lines) > 0 else 0), end = (int(lines[1]) - 1 if len(lines) > 1 else number)))

    def readyPrint(self):
        if(os.stat(name).st_size == 0):
            print("Пустой файл")
        elif(len(filterVacansie[0]) > 0 and len(filterVacansie) == 1):
            print("Формат ввода некорректен")
        elif(filterVacansie[0] not in fieldToEng.keys() and filterVacansie[0] != ""):
            print("Параметр поиска некорректен")
        elif(sortingParameter not in fieldToEng.keys() and sortingParameter != ""):
            print("Параметр сортировки некорректен")
        elif(reverseOrder != "Да" and reverseOrder != "Нет" and reverseOrder != ""):
            print("Порядок сортировки задан некорректно")
        else:
            a, b = self.data.сsv_reader(name)
            if len(b) == 0 or len(a) == 0:
                print("Нет данных")
            else:
                self.print_vacancies(self.data.csv_ﬁler(a, b), fieldToRus)  



class Report:
    def __init__(self):
        self.dictionary_salary_levels = {}
        self.dictionary_number_vacancies = {}
        self.dictionary_salary_levels_profession = {}
        self.dictionary_number_vacancies_profession = {}
        self.dictionary_level_salaries_cities = {}
        self.dictionary_vacancy_rate_city = {}

    def generate_graphics(self):
        plt.rcParams.update({'font.size': 8})
        #1 ГРАФИК
        labelsLevel = self.dictionary_salary_levels.keys()
        middle_salary = self.dictionary_salary_levels.values()
        middle_salary_profession = self.dictionary_salary_levels_profession.values()

        x = np.arange(len(labelsLevel))  
        width = 0.35 

        ax = plt.subplot(221)
        ax.bar(x - width/2, middle_salary, width, label='средняя з/п')
        ax.bar(x + width/2, middle_salary_profession, width, label='з/п ' + profession)
        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(x, labelsLevel, rotation = 90)
        plt.grid(axis='y')
        ax.legend()
        #2 ГРАФИК
        labelsCount = self.dictionary_number_vacancies.keys()
        middle_count = self.dictionary_number_vacancies.values()
        middle_count_profession = self.dictionary_number_vacancies_profession.values()

        x = np.arange(len(labelsCount))  
        width = 0.35 

        ax = plt.subplot(222)
        ax.bar(x - width/2, middle_count, width, label='Количества вакансий')
        ax.bar(x + width/2, middle_count_profession, width, label='Количества вакансий \n' + profession)
        ax.set_title('Количество вакансий по годам')
        ax.set_xticks(x, labelsCount, rotation = 90)
        plt.grid(axis='y')
        ax.legend()
        #3 ГРАФИК
        ax = plt.subplot(223)
        ax.set_title('Уровень зарплат по городам')
        plt.rcParams.update({'font.size': 6})
        city = self.dictionary_level_salaries_cities.keys()
        y_pos = np.arange(len(city))
        performance = self.dictionary_level_salaries_cities.values()

        ax.barh(y_pos, performance,  align = 'center')
        ax.set_yticks(y_pos, labels = city)
        ax.invert_yaxis() 
        plt.grid(axis='x')

        #4 ГРАФИК   
        ax = plt.subplot(224)
        plt.rcParams.update({'font.size': 8})
        ax.set_title('Доля вакансий по городам')
        plt.rcParams.update({'font.size': 6})
        ax.pie(list(self.dictionary_vacancy_rate_city.values()) + [1 - sum(self.dictionary_vacancy_rate_city.values())], labels = list(self.dictionary_vacancy_rate_city.keys()) + ["Другие"])

        plt.subplots_adjust(wspace = 0.5, hspace = 0.5)
        plt.savefig("Graphics", dpi = 200, bbox_inches='tight')

    def set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)    

    def set_bold(self, ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold = True)

    def width(self, ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) + 0.5
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate_excel(self):
        wb = Workbook()
        #1 СТРАНИЦА
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        ws1.append(["Год", "Средняя зарплата", "Средняя зарплата - " + profession, "Количество вакансий", "Количество вакансий - " + profession])
        
        countIndexA = 2
        for year in self.dictionary_salary_levels.keys():
            ws1['A' + str(countIndexA)] = year
            countIndexA += 1
        countIndexB = 2
        for value in self.dictionary_salary_levels.values():
            ws1['B' + str(countIndexB)] = value
            countIndexB += 1
        countIndexC = 2
        for value in self.dictionary_salary_levels_profession.values():
            ws1['C' + str(countIndexC)] = value
            countIndexC += 1
        countIndexD = 2
        for value in self.dictionary_number_vacancies.values():
            ws1['D' + str(countIndexD)] = value
            countIndexD += 1
        countIndexE = 2
        for value in self.dictionary_number_vacancies_profession.values():
            ws1['E' + str(countIndexE)] = value
            countIndexE += 1
        
        self.width(ws1)
        self.set_border(ws1, 'A1:E' + str(1 + len(self.dictionary_salary_levels.keys()))) 
        self.set_bold(ws1, 'A1:E1') 
        #2 СТРАНИЦА
        ws2 = wb.create_sheet("Статистика по городам")
        ws2.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        
        countWS2IndexA = 2
        for city in self.dictionary_level_salaries_cities.keys():
            ws2['A' + str(countWS2IndexA)] = city
            countWS2IndexA += 1
        countWS2IndexB = 2
        for value in self.dictionary_level_salaries_cities.values():
            ws2['B' + str(countWS2IndexB)] = value
            countWS2IndexB += 1
        countWS2IndexD = 2
        for city in self.dictionary_vacancy_rate_city.keys():
            ws2['D' + str(countWS2IndexD)] = city
            countWS2IndexD += 1
        countWS2IndexE = 2
        for value in self.dictionary_vacancy_rate_city.values():
            ws2['E' + str(countWS2IndexE)] = value
            ws2['E' + str(countWS2IndexE)].number_format = '0.00%'
            countWS2IndexE += 1

        self.set_border(ws2, 'A1:B' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_border(ws2, 'D1:E' + str(1 + len(self.dictionary_level_salaries_cities.keys()))) 
        self.set_bold(ws2, 'A1:E1') 
        self.width(ws2)
        wb.save('report.xlsx')

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")
        
        pdf_template = template.render({"name": profession})
        config = pdfkit.configuration(wkhtmltopdf = r'D:\Учеба\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdf_template = pdf_template.replace("path", os.path.abspath(os.curdir))
        options = {'enable-local-file-access': None}

        readyHtmlTable = self.generate_table()
        pdf_template = pdf_template.replace("$table1;", readyHtmlTable)
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def generate_table(self):
        #1 ТАБЛИЦА
        readyHtmlTable = "<table class='main_table'><tr><th>Год</th><th>Средняя зарплата</th><th>Средняя зарплата - "
        readyHtmlTable += profession + "</th><th>Количество вакансий</th><th>Количество вакансий - " + profession + "</th></tr>"
        for index in range(len(list(self.dictionary_salary_levels.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_salary_levels_profession.values())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_number_vacancies_profession.values())[index]) + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table><h1>Статистика по городам</h1>"
        #2 ТАБЛИЦА
        readyHtmlTable += "<table class='tableCity1'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.dictionary_level_salaries_cities.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_level_salaries_cities.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(list(self.dictionary_level_salaries_cities.values())[index]) + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table>"
        #3 ТАБЛИЦА
        readyHtmlTable += "<table class='tableCity2'><tr><th>Город</th><th>Уровень зарплат</th>"
        for index in range(len(list(self.dictionary_vacancy_rate_city.keys()))):
            readyHtmlTable += "<tr>"
            readyHtmlTable += ("<td>" + str(list(self.dictionary_vacancy_rate_city.keys())[index]) + "</td>")
            readyHtmlTable += ("<td>" + str(round(list(self.dictionary_vacancy_rate_city.values())[index] * 100, 2)) + "%" + "</td>")
            readyHtmlTable += "</tr>"
        readyHtmlTable += "</tr></table>"
        return readyHtmlTable
if(choice == "Статистика"):  
    printer = DataSet()
    printer.readyPrint()
if(choice == "Вакансии"):
    printer = InputConnect()
    printer.readyPrint()